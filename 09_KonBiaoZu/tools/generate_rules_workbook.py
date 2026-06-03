from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:\Users\Administrator\Documents\New project 2\KonBiaoZu")
CSV_PATH = ROOT / "data" / "KonBiaoZuRules.csv"
XLSX_PATH = ROOT / "data" / "KonBiaoZuRules.xlsx"

CSV_HEADERS = [
    "标注类型",
    "子类型",
    "规格",
    "螺纹/型号",
    "标准长度",
    "底孔",
    "显示标注",
    "标准注释",
    "备注",
]

TYPE_ORDER = [
    "孔标注",
    "螺牙标注",
    "焊接螺母标注",
    "压铆螺母标注",
    "压铆螺钉标注",
    "压铆螺母柱标注",
    "沉孔标注",
]

SOURCE_ROWS = [
    {"type": "螺牙标注", "subtype": "公制粗牙", "size": "M3", "thread": "M3x0.5", "length": "", "hole": "2.50", "display": "攻牙 {规格}", "comment": "{螺纹} 底孔 Φ{底孔}", "note": "标准公制粗牙底孔"},
    {"type": "螺牙标注", "subtype": "公制粗牙", "size": "M4", "thread": "M4x0.7", "length": "", "hole": "3.30", "display": "攻牙 {规格}", "comment": "{螺纹} 底孔 Φ{底孔}", "note": "标准公制粗牙底孔"},
    {"type": "螺牙标注", "subtype": "公制粗牙", "size": "M5", "thread": "M5x0.8", "length": "", "hole": "4.20", "display": "攻牙 {规格}", "comment": "{螺纹} 底孔 Φ{底孔}", "note": "标准公制粗牙底孔"},
    {"type": "螺牙标注", "subtype": "公制粗牙", "size": "M6", "thread": "M6x1.0", "length": "", "hole": "5.00", "display": "攻牙 {规格}", "comment": "{螺纹} 底孔 Φ{底孔}", "note": "标准公制粗牙底孔"},
    {"type": "焊接螺母标注", "subtype": "WN", "size": "M3", "thread": "WN-M3", "length": "", "hole": "4.39", "display": "焊接螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "碳钢系列"},
    {"type": "焊接螺母标注", "subtype": "WN", "size": "M4", "thread": "WN-M4", "length": "", "hole": "5.53", "display": "焊接螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "碳钢系列"},
    {"type": "焊接螺母标注", "subtype": "WN", "size": "M5", "thread": "WN-M5", "length": "", "hole": "6.35", "display": "焊接螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "碳钢系列"},
    {"type": "焊接螺母标注", "subtype": "WN", "size": "M6", "thread": "WN-M6", "length": "", "hole": "8.04", "display": "焊接螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "碳钢系列"},
    {"type": "焊接螺母标注", "subtype": "WNS", "size": "M3", "thread": "WNS-M3", "length": "", "hole": "4.39", "display": "不锈钢焊接螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "不锈钢系列"},
    {"type": "焊接螺母标注", "subtype": "WNS", "size": "M4", "thread": "WNS-M4", "length": "", "hole": "5.53", "display": "不锈钢焊接螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "不锈钢系列"},
    {"type": "焊接螺母标注", "subtype": "WNS", "size": "M5", "thread": "WNS-M5", "length": "", "hole": "6.35", "display": "不锈钢焊接螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "不锈钢系列"},
    {"type": "焊接螺母标注", "subtype": "WNS", "size": "M6", "thread": "WNS-M6", "length": "", "hole": "8.04", "display": "不锈钢焊接螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "不锈钢系列"},
    {"type": "压铆螺母标注", "subtype": "S", "size": "M3", "thread": "S-M3", "length": "", "hole": "4.22", "display": "压铆螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "碳钢系列"},
    {"type": "压铆螺母标注", "subtype": "S", "size": "M4", "thread": "S-M4", "length": "", "hole": "5.41", "display": "压铆螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "碳钢系列"},
    {"type": "压铆螺母标注", "subtype": "S", "size": "M5", "thread": "S-M5", "length": "", "hole": "6.35", "display": "压铆螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "碳钢系列"},
    {"type": "压铆螺母标注", "subtype": "S", "size": "M6", "thread": "S-M6", "length": "", "hole": "8.75", "display": "压铆螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "碳钢系列"},
    {"type": "压铆螺母标注", "subtype": "CLS", "size": "M3", "thread": "CLS-M3", "length": "", "hole": "4.22", "display": "不锈钢压铆螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "不锈钢系列"},
    {"type": "压铆螺母标注", "subtype": "CLS", "size": "M4", "thread": "CLS-M4", "length": "", "hole": "5.41", "display": "不锈钢压铆螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "不锈钢系列"},
    {"type": "压铆螺母标注", "subtype": "CLS", "size": "M5", "thread": "CLS-M5", "length": "", "hole": "6.35", "display": "不锈钢压铆螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "不锈钢系列"},
    {"type": "压铆螺母标注", "subtype": "CLS", "size": "M6", "thread": "CLS-M6", "length": "", "hole": "8.75", "display": "不锈钢压铆螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "不锈钢系列"},
    {"type": "压铆螺母标注", "subtype": "SP", "size": "M3", "thread": "SP-M3", "length": "", "hole": "4.22", "display": "高强压铆螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "高强系列，待补充校核"},
    {"type": "压铆螺母标注", "subtype": "SP", "size": "M4", "thread": "SP-M4", "length": "", "hole": "5.41", "display": "高强压铆螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "高强系列，待补充校核"},
    {"type": "压铆螺母标注", "subtype": "SP", "size": "M5", "thread": "SP-M5", "length": "", "hole": "6.35", "display": "高强压铆螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "高强系列，待补充校核"},
    {"type": "压铆螺母标注", "subtype": "SP", "size": "M6", "thread": "SP-M6", "length": "", "hole": "8.75", "display": "高强压铆螺母 {规格}", "comment": "PEM {螺纹}，底孔 Φ{底孔}", "note": "高强系列，待补充校核"},
    {"type": "压铆螺钉标注", "subtype": "FH", "size": "M3", "thread": "FH-M3", "length": "按输入", "hole": "4.22", "display": "压铆螺钉 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺钉标注", "subtype": "FH", "size": "M4", "thread": "FH-M4", "length": "按输入", "hole": "5.41", "display": "压铆螺钉 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺钉标注", "subtype": "FH", "size": "M5", "thread": "FH-M5", "length": "按输入", "hole": "6.35", "display": "压铆螺钉 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺钉标注", "subtype": "FH", "size": "M6", "thread": "FH-M6", "length": "按输入", "hole": "8.75", "display": "压铆螺钉 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺钉标注", "subtype": "FHS", "size": "M3", "thread": "FHS-M3", "length": "按输入", "hole": "4.22", "display": "不锈钢压铆螺钉 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺钉标注", "subtype": "FHS", "size": "M4", "thread": "FHS-M4", "length": "按输入", "hole": "5.41", "display": "不锈钢压铆螺钉 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺钉标注", "subtype": "FHS", "size": "M5", "thread": "FHS-M5", "length": "按输入", "hole": "6.35", "display": "不锈钢压铆螺钉 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺钉标注", "subtype": "FHS", "size": "M6", "thread": "FHS-M6", "length": "按输入", "hole": "8.75", "display": "不锈钢压铆螺钉 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺母柱标注", "subtype": "SO", "size": "M3", "thread": "SO-M3", "length": "按输入", "hole": "5.40", "display": "通孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺母柱标注", "subtype": "SO", "size": "M4", "thread": "SO-M4", "length": "按输入", "hole": "5.40", "display": "通孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺母柱标注", "subtype": "SO", "size": "M5", "thread": "SO-M5", "length": "按输入", "hole": "7.10", "display": "通孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺母柱标注", "subtype": "SO", "size": "M6", "thread": "SO-M6", "length": "按输入", "hole": "9.10", "display": "通孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺母柱标注", "subtype": "SOS", "size": "M3", "thread": "SOS-M3", "length": "按输入", "hole": "5.40", "display": "不锈钢通孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺母柱标注", "subtype": "SOS", "size": "M4", "thread": "SOS-M4", "length": "按输入", "hole": "5.40", "display": "不锈钢通孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺母柱标注", "subtype": "SOS", "size": "M5", "thread": "SOS-M5", "length": "按输入", "hole": "7.10", "display": "不锈钢通孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺母柱标注", "subtype": "SOS", "size": "M6", "thread": "SOS-M6", "length": "按输入", "hole": "9.10", "display": "不锈钢通孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入"},
    {"type": "压铆螺母柱标注", "subtype": "BSO", "size": "M3", "thread": "BSO-M3", "length": "按输入", "hole": "5.40", "display": "盲孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入，底孔待你们标准复核"},
    {"type": "压铆螺母柱标注", "subtype": "BSO", "size": "M4", "thread": "BSO-M4", "length": "按输入", "hole": "5.40", "display": "盲孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入，底孔待你们标准复核"},
    {"type": "压铆螺母柱标注", "subtype": "BSO", "size": "M5", "thread": "BSO-M5", "length": "按输入", "hole": "7.10", "display": "盲孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入，底孔待你们标准复核"},
    {"type": "压铆螺母柱标注", "subtype": "BSO", "size": "M6", "thread": "BSO-M6", "length": "按输入", "hole": "9.10", "display": "盲孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入，底孔待你们标准复核"},
    {"type": "压铆螺母柱标注", "subtype": "BSOS", "size": "M3", "thread": "BSOS-M3", "length": "按输入", "hole": "5.40", "display": "不锈钢盲孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入，底孔待你们标准复核"},
    {"type": "压铆螺母柱标注", "subtype": "BSOS", "size": "M4", "thread": "BSOS-M4", "length": "按输入", "hole": "5.40", "display": "不锈钢盲孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入，底孔待你们标准复核"},
    {"type": "压铆螺母柱标注", "subtype": "BSOS", "size": "M5", "thread": "BSOS-M5", "length": "按输入", "hole": "7.10", "display": "不锈钢盲孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入，底孔待你们标准复核"},
    {"type": "压铆螺母柱标注", "subtype": "BSOS", "size": "M6", "thread": "BSOS-M6", "length": "按输入", "hole": "9.10", "display": "不锈钢盲孔螺母柱 {规格}", "comment": "PEM {螺纹}-L{长度}，安装孔 Φ{底孔}", "note": "长度由对话框输入，底孔待你们标准复核"},
    {"type": "沉孔标注", "subtype": "标准沉孔", "size": "Φ4", "thread": "标准沉孔-Φ4", "length": "", "hole": "4.00", "display": "沉孔 {规格}", "comment": "M3螺丝沉孔", "note": "沉孔按小孔孔径区间判定：n.1~n+1.0 对应 Mn"},
    {"type": "沉孔标注", "subtype": "标准沉孔", "size": "Φ5", "thread": "标准沉孔-Φ5", "length": "", "hole": "5.00", "display": "沉孔 {规格}", "comment": "M4螺丝沉孔", "note": "沉孔按小孔孔径区间判定：n.1~n+1.0 对应 Mn"},
    {"type": "沉孔标注", "subtype": "标准沉孔", "size": "Φ5.1", "thread": "标准沉孔-Φ5.1", "length": "", "hole": "5.10", "display": "沉孔 {规格}", "comment": "M5螺丝沉孔", "note": "沉孔按小孔孔径区间判定：n.1~n+1.0 对应 Mn"},
    {"type": "沉孔标注", "subtype": "标准沉孔", "size": "Φ6", "thread": "标准沉孔-Φ6", "length": "", "hole": "6.00", "display": "沉孔 {规格}", "comment": "M5螺丝沉孔", "note": "沉孔按小孔孔径区间判定：n.1~n+1.0 对应 Mn"},
    {"type": "沉孔标注", "subtype": "标准沉孔", "size": "Φ7", "thread": "标准沉孔-Φ7", "length": "", "hole": "7.00", "display": "沉孔 {规格}", "comment": "M6螺丝沉孔", "note": "沉孔按小孔孔径区间判定：n.1~n+1.0 对应 Mn"},
    {"type": "沉孔标注", "subtype": "标准沉孔", "size": "Φ8", "thread": "标准沉孔-Φ8", "length": "", "hole": "8.00", "display": "沉孔 {规格}", "comment": "M7螺丝沉孔", "note": "沉孔按小孔孔径区间判定：n.1~n+1.0 对应 Mn"},
    {"type": "沉孔标注", "subtype": "标准沉孔", "size": "Φ9", "thread": "标准沉孔-Φ9", "length": "", "hole": "9.00", "display": "沉孔 {规格}", "comment": "M8螺丝沉孔", "note": "沉孔按小孔孔径区间判定：n.1~n+1.0 对应 Mn"},
    {"type": "沉孔标注", "subtype": "标准沉孔", "size": "Φ10", "thread": "标准沉孔-Φ10", "length": "", "hole": "10.00", "display": "沉孔 {规格}", "comment": "M9螺丝沉孔", "note": "沉孔按小孔孔径区间判定：n.1~n+1.0 对应 Mn"},
    {"type": "沉孔标注", "subtype": "标准沉孔", "size": "Φ12", "thread": "标准沉孔-Φ12", "length": "", "hole": "12.00", "display": "沉孔 {规格}", "comment": "M11螺丝沉孔", "note": "沉孔按小孔孔径区间判定：n.1~n+1.0 对应 Mn"},
    {"type": "沉孔标注", "subtype": "标准沉孔", "size": "Φ14", "thread": "标准沉孔-Φ14", "length": "", "hole": "14.00", "display": "沉孔 {规格}", "comment": "M13螺丝沉孔", "note": "沉孔按小孔孔径区间判定：n.1~n+1.0 对应 Mn"},
]

PLACEHOLDER_SHEETS = {
    "孔标注": {
        "subtypes": ["普通孔"],
        "display": {"普通孔": "孔 {规格}"},
        "comment": {"普通孔": "普通孔，孔径 Φ{底孔}"},
        "length": {"普通孔": ""},
        "note": {"普通孔": "后续按企业标准补充"},
        "rows": [],
    },
    "沉孔标注": {
        "subtypes": ["标准沉孔"],
        "display": {"标准沉孔": "沉孔 {规格}"},
        "comment": {"标准沉孔": "沉孔 {规格}，底孔 Φ{底孔}"},
        "length": {"标准沉孔": ""},
        "note": {"标准沉孔": "后续按企业标准补充"},
        "rows": [],
    },
}


def render_template(template: str, row: dict[str, str]) -> str:
    length_value = row["length"] if row["length"] else ""
    return (
        template.replace("{子类型}", row["subtype"])
        .replace("{规格}", row["size"])
        .replace("{螺纹}", row["thread"])
        .replace("{底孔}", row["hole"])
        .replace("{长度}", length_value)
    )


def build_sheet_models() -> dict[str, dict[str, object]]:
    grouped_rows: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in SOURCE_ROWS:
        grouped_rows[row["type"]].append(row)

    sheet_models: dict[str, dict[str, object]] = {}
    for sheet_name in TYPE_ORDER:
        if sheet_name in PLACEHOLDER_SHEETS:
            sheet_models[sheet_name] = PLACEHOLDER_SHEETS[sheet_name]
            continue

        rows = grouped_rows.get(sheet_name, [])
        subtypes: list[str] = []
        display_map: dict[str, str] = {}
        comment_map: dict[str, str] = {}
        length_map: dict[str, str] = {}
        note_map: dict[str, str] = {}
        hole_rows: dict[str, dict[str, str]] = defaultdict(dict)

        for row in rows:
            subtype = row["subtype"]
            if subtype not in subtypes:
                subtypes.append(subtype)
                display_map[subtype] = row["display"]
                comment_map[subtype] = row["comment"]
                length_map[subtype] = row["length"]
                note_map[subtype] = row["note"]
            current_value = hole_rows[row["hole"]].get(subtype, "")
            if not current_value:
                hole_rows[row["hole"]][subtype] = row["size"]
            elif row["size"] not in current_value.split(" / "):
                hole_rows[row["hole"]][subtype] = current_value + " / " + row["size"]

        ordered_holes = sorted(hole_rows.keys(), key=lambda value: float(value))
        matrix_rows = [(hole, hole_rows[hole]) for hole in ordered_holes]
        sheet_models[sheet_name] = {
            "subtypes": subtypes,
            "display": display_map,
            "comment": comment_map,
            "length": length_map,
            "note": note_map,
            "rows": matrix_rows,
        }

    return sheet_models


def create_type_sheet(workbook: Workbook, sheet_name: str, model: dict[str, object]) -> None:
    ws = workbook.create_sheet(sheet_name)
    subtypes: list[str] = model["subtypes"]  # type: ignore[assignment]
    display_map: dict[str, str] = model["display"]  # type: ignore[assignment]
    comment_map: dict[str, str] = model["comment"]  # type: ignore[assignment]
    length_map: dict[str, str] = model["length"]  # type: ignore[assignment]
    note_map: dict[str, str] = model["note"]  # type: ignore[assignment]
    matrix_rows: list[tuple[str, dict[str, str]]] = model["rows"]  # type: ignore[assignment]

    title_fill = PatternFill("solid", fgColor="0F4C81")
    label_fill = PatternFill("solid", fgColor="D9EAF7")
    subtype_fill = PatternFill("solid", fgColor="F6C85F")
    zebra_fill = PatternFill("solid", fgColor="F7FAFC")
    thin = Side(style="thin", color="D0D7DE")

    end_col = get_column_letter(max(2, len(subtypes) + 1))
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = f"{sheet_name} 配置表"
    ws["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    labels = ["类别", "标注文本", "注释模板", "长度规则", "说明", "底孔"]
    for row_index, label in enumerate(labels, start=2):
        cell = ws.cell(row=row_index, column=1, value=label)
        cell.font = Font(name="Microsoft YaHei", bold=True)
        cell.fill = label_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for offset, subtype in enumerate(subtypes, start=2):
        cells = {
            2: subtype,
            3: display_map.get(subtype, ""),
            4: comment_map.get(subtype, ""),
            5: length_map.get(subtype, ""),
            6: note_map.get(subtype, ""),
        }
        for row_index, value in cells.items():
            cell = ws.cell(row=row_index, column=offset, value=value)
            cell.font = Font(name="Microsoft YaHei", bold=(row_index == 2))
            if row_index == 2:
                cell.fill = subtype_fill
            cell.alignment = Alignment(horizontal="center" if row_index == 2 else "left", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    data_start = 7
    for row_offset, (hole_value, mapping) in enumerate(matrix_rows):
        excel_row = data_start + row_offset
        hole_cell = ws.cell(row=excel_row, column=1, value=hole_value)
        hole_cell.font = Font(name="Microsoft YaHei")
        hole_cell.alignment = Alignment(horizontal="center", vertical="center")
        hole_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if row_offset % 2 == 0:
            hole_cell.fill = zebra_fill

        for col_offset, subtype in enumerate(subtypes, start=2):
            cell = ws.cell(row=excel_row, column=col_offset, value=mapping.get(subtype, ""))
            cell.font = Font(name="Microsoft YaHei")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if row_offset % 2 == 0:
                cell.fill = zebra_fill

    ws.freeze_panes = "B7"
    ws.column_dimensions["A"].width = 18
    for index in range(2, len(subtypes) + 2):
        ws.column_dimensions[get_column_letter(index)].width = 24

    tip_row = max(data_start + len(matrix_rows) + 2, 10)
    ws[f"A{tip_row}"] = "占位符说明"
    ws[f"A{tip_row}"].font = Font(name="Microsoft YaHei", size=11, bold=True)
    ws[f"A{tip_row + 1}"] = "标注文本和注释模板支持：{规格} {螺纹} {底孔} {长度} {子类型}"
    ws[f"A{tip_row + 2}"] = "示例：把 FHS 的标注文本改成“FHS标注 {规格}”或直接写“FHS标注”。"
    ws[f"A{tip_row + 3}"] = "程序后续建议直接按单元格读取；当前 CSV 由本脚本同步导出。"
    ws.column_dimensions["A"].width = 26


def create_usage_sheet(workbook: Workbook) -> None:
    ws = workbook.create_sheet("使用说明")
    ws["A1"] = "KonBiaoZu 规则表使用说明"
    ws["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True)
    instructions = [
        "1. 每张表只对应一种标注类型，不再保留旧的总表/映射表/文本自定义表。",
        "2. 第 2 行是子类型，第 3 行是可直接修改的标注文本模板，第 4 行是可直接修改的注释模板。",
        "3. 第 7 行开始是底孔和规格映射，程序按“主类型 + 子类型 + 底孔”匹配规格。",
        "4. 如果要把 FHS 默认文字改成企业叫法，直接在“压铆螺钉标注”页把 FHS 列的“标注文本”单元格改掉即可。",
        "5. 现在脚本会同时生成 CSV 镜像给插件使用；后面如果改成程序直接读 xlsx，可以保留这套表结构不变。",
    ]
    for index, text in enumerate(instructions, start=3):
        ws[f"A{index}"] = text
    ws.column_dimensions["A"].width = 110


def export_csv() -> None:
    rows = []
    for row in SOURCE_ROWS:
        rows.append(
            [
                row["type"],
                row["subtype"],
                row["size"],
                row["thread"],
                row["length"],
                row["hole"],
                render_template(row["display"], row),
                render_template(row["comment"], row),
                row["note"],
            ]
        )

    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(CSV_HEADERS)
        writer.writerows(rows)


def main() -> None:
    export_csv()
    models = build_sheet_models()
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in TYPE_ORDER:
        create_type_sheet(workbook, sheet_name, models[sheet_name])
    create_usage_sheet(workbook)
    workbook.save(XLSX_PATH)
    print("ok")


if __name__ == "__main__":
    main()
